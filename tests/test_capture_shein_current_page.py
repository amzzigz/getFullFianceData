from pathlib import Path

from openpyxl import load_workbook

from scripts.capture_shein_current_page import append_rows, repair_current_page_images


def test_append_rows_deduplicates_by_goods_id(tmp_path: Path) -> None:
    output = tmp_path / "products.xlsx"
    payload = {
        "title": "Search grdr sweater | SHEIN USA",
        "url": "https://us.shein.com/pdsearch/grdr%20sweater/?page=1",
        "rows": [
            {
                "goods_id": "1",
                "name": "A",
                "image": "https://img.ltwebstatic.com/a.jpg",
                "sales": "",
                "price": "$1.00",
                "original_price": "$2.00",
                "link": "https://us.shein.com/A-p-1.html",
            },
            {
                "goods_id": "1",
                "name": "A duplicate",
                "image": "https://img.ltwebstatic.com/a2.jpg",
                "sales": "",
                "price": "$1.00",
                "original_price": "$2.00",
                "link": "https://us.shein.com/A-p-1.html",
            },
        ],
    }

    added, found = append_rows(output, payload)

    wb = load_workbook(output)
    ws = wb["商品"]
    assert found == 2
    assert added == 1
    assert ws.max_row == 2
    assert ws["D2"].value == "1"


def test_repair_current_page_images_only_updates_matching_goods_id(tmp_path: Path) -> None:
    output = tmp_path / "products.xlsx"
    append_rows(
        output,
        {
            "title": "Page",
            "url": "https://us.shein.com/page=1",
            "rows": [
                {"goods_id": "1", "name": "A", "image": "wrong", "sales": "5 sold", "price": "$1", "original_price": "$2", "link": "a"},
                {"goods_id": "2", "name": "B", "image": "wrong", "sales": "6 sold", "price": "$3", "original_price": "$4", "link": "b"},
            ],
        },
    )

    updated, found = repair_current_page_images(
        output,
        {
            "rows": [
                {"goods_id": "1", "image": "https://img.ltwebstatic.com/real-main.webp"},
                {"goods_id": "2", "image": "https://sc.ltwebstatic.com/she_dist/images/bg-grey.png"},
            ]
        },
    )

    ws = load_workbook(output)["商品"]
    assert (updated, found) == (1, 1)
    assert ws["E2"].value == "A"
    assert ws["F2"].value == "https://img.ltwebstatic.com/real-main.webp"
    assert ws["F3"].value == "wrong"

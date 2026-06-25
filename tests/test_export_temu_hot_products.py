from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from scripts.export_temu_hot_products import (
    HEADERS,
    collect_hot_products,
    find_hot_result,
    open_global_sales_page,
    parse_hot_rows,
    write_excel,
)


def hot_result() -> dict:
    return {
        "total": 1,
        "subOrderList": [
            {
                "productName": "商品名",
                "category": "女装针织衫",
                "productSkcId": 123456,
                "skcExtCode": "STYLE-RED",
                "hotTag": True,
                "skuQuantityDetailList": [
                    {"supplierPrice": 3500, "skuExtCode": "STYLE-RED-S"},
                    {"supplierPrice": 3500, "skuExtCode": "STYLE-RED-M"},
                ],
            }
        ],
    }


def test_find_hot_result_uses_last_hot_tag_response() -> None:
    normal = {"result": {"total": 9, "subOrderList": []}}
    hot = {"result": hot_result()}
    har = {
        "log": {
            "entries": [
                {
                    "request": {
                        "url": "https://agentseller.temu.com/mms/venom/api/supplier/sales/management/listOverall",
                        "postData": {"text": '{"pageNo":1,"hotTag":false}'},
                    },
                    "response": {"content": {"text": str(normal)}},
                },
                {
                    "request": {
                        "url": "https://agentseller.temu.com/mms/venom/api/supplier/sales/management/listOverall",
                        "postData": {"text": '{"pageNo":1,"hotTag":true}'},
                    },
                    "response": {"content": {"text": '{"result":{"total":1,"subOrderList":[]}}'}},
                },
                {
                    "request": {
                        "url": "https://agentseller.temu.com/mms/venom/api/supplier/sales/management/listOverall",
                        "postData": {"text": '{"pageNo":1,"pageSize":40,"hotTag":true}'},
                    },
                    "response": {"content": {"text": '{"result":{"total":3,"subOrderList":[{"productName":"最后一次"}]}}'}},
                },
            ]
        }
    }

    result = find_hot_result(har)

    assert result["total"] == 3
    assert result["subOrderList"][0]["productName"] == "最后一次"


def test_parse_hot_rows_returns_one_row_per_skc_and_only_requested_fields() -> None:
    rows = parse_hot_rows("B23/B25/B26-主账号-YF", "MinimalKnit", hot_result())

    assert rows == [
        (
            "B23/B25/B26-主账号-YF",
            "MinimalKnit",
            "商品名",
            "女装针织衫",
            "123456",
            "STYLE-RED",
            35,
        )
    ]


def test_parse_hot_rows_joins_distinct_prices() -> None:
    result = hot_result()
    result["subOrderList"][0]["skuQuantityDetailList"][1]["supplierPrice"] = 3550

    rows = parse_hot_rows("账号", "店铺", result)

    assert rows[0][-1] == "35/35.5"


def test_write_excel_has_exact_requested_headers(tmp_path: Path) -> None:
    output = tmp_path / "hot.xlsx"
    write_excel(parse_hot_rows("账号", "店铺", hot_result()), output)

    ws = load_workbook(output).active

    assert [cell.value for cell in ws[1]] == HEADERS
    assert ws.max_column == 7
    assert ws.max_row == 2


def test_collect_hot_products_fetches_all_pages() -> None:
    payloads: list[dict] = []

    def post_json(_page, _url, payload, _mall_id):
        payloads.append(payload)
        return {
            "success": True,
            "errorCode": 1000000,
            "result": {
                "total": 3,
                "subOrderList": [{"productSkcId": payload["pageNo"]}],
            },
        }

    records = collect_hot_products(object(), 88, post_json=post_json, page_size=2)

    assert records == [{"productSkcId": 1}, {"productSkcId": 2}]
    assert payloads == [
        {"pageNo": 1, "pageSize": 2, "isLack": 0, "hotTag": True},
        {"pageNo": 2, "pageSize": 2, "isLack": 0, "hotTag": True},
    ]


def test_open_global_sales_page_uses_official_link_and_confirms_agent_session(monkeypatch) -> None:
    navigations: list[str] = []
    posts: list[tuple[str, object]] = []
    handler_calls: list[str] = []

    class FakePage:
        url = "https://seller.kuajingmaihuo.com/"

        def get(self, url: str) -> None:
            navigations.append(url)
            self.url = "https://agentseller.temu.com/stock-entry"

    class FakeHelper:
        def _log(self, _message: str) -> None:
            pass

        def _handle_click_for_platform(self, page, _platform, lower_url, _log, _browser):
            handler_calls.append(lower_url)
            return page

    class FakeContext:
        page = FakePage()
        helper = FakeHelper()
        browser = object()

    monkeypatch.setattr("scripts.export_temu_hot_products.ensure_seller_page", lambda _ctx: None)
    monkeypatch.setattr("scripts.export_temu_hot_products.set_seller_mall_context", lambda _page, _mall_id: None)
    monkeypatch.setattr("scripts.export_temu_hot_products.time.sleep", lambda _seconds: None)

    def fake_post(_page, url, _payload, mall_id):
        posts.append((url, mall_id))
        return {"success": True, "errorCode": 1000000, "result": {}}

    monkeypatch.setattr("scripts.export_temu_hot_products.browser_post_json", fake_post)

    open_global_sales_page(FakeContext(), 634418228348855, timeout_seconds=1)

    assert navigations[0].startswith("https://seller.kuajingmaihuo.com/link-agent-seller?region=1&targetUrl=")
    assert "stock-entry" in navigations[0]
    assert posts == [("https://agentseller.temu.com/api/seller/auth/userInfo", 634418228348855)]
    assert handler_calls == []

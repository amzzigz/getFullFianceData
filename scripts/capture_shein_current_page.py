from __future__ import annotations

import argparse
import json
import os
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.error import URLError
from urllib.request import urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright


os.environ.setdefault("NODE_NO_WARNINGS", "1")

DEFAULT_OUTPUT = Path.home() / "Desktop" / "shein_current_page_products.xlsx"
HEADERS = [
    "抓取时间",
    "页面标题",
    "当前页URL",
    "goods_id",
    "商品名",
    "主图",
    "销量",
    "价格",
    "原价",
    "商品链接",
]


EXTRACT_JS = r"""
() => {
  const absolute = (url) => {
    if (!url) return "";
    if (url.startsWith("//")) return "https:" + url;
    try { return new URL(url, location.href).href; } catch (_) { return url; }
  };
  const clean = (text) => (text || "").replace(/\s+/g, " ").trim();
  const productImage = (img) => {
    if (!img) return "";
    const candidates = [img.getAttribute("data-src"), img.getAttribute("data-original"), img.getAttribute("src")];
    for (const candidate of candidates) {
      const url = absolute(candidate || "");
      if (url.includes("ltwebstatic.com") && !url.includes("bg-grey")) return url;
    }
    return "";
  };
  const goodsIdFromHref = (href) => {
    const match = String(href || "").match(/-p-(\d+)(?:[^\d]|$)/i) || String(href || "").match(/[?&]goods_id=(\d+)/i);
    return match ? match[1] : "";
  };
  const priceMatches = (text) => Array.from(String(text || "").matchAll(/\$\s*\d+(?:\.\d{1,2})?/g)).map((m) => clean(m[0]));
  const salesFromText = (text) => {
    const match = String(text || "").match(/(?:\d+(?:\.\d+)?\s*[Kk+]?\s*(?:sold|销量|已售|售出))|(?:(?:sold|销量|已售|售出)\s*\d+(?:\.\d+)?\s*[Kk+]?)/i);
    return match ? clean(match[0]) : "";
  };
  const titleFrom = (anchor, card) => {
    const img = card.querySelector("img[alt], img[title]") || anchor.querySelector("img[alt], img[title]");
    const candidates = [
      anchor.getAttribute("aria-label"),
      anchor.getAttribute("title"),
      img && img.getAttribute("alt"),
      img && img.getAttribute("title"),
    ].filter(Boolean);
    for (const item of candidates) {
      const value = clean(item);
      if (value && !/^\$/.test(value)) return value;
    }
    const lines = clean(card.innerText).split(/ (?=\$)|\n/).map(clean).filter(Boolean);
    return lines.find((line) => !line.includes("$") && !/sold|销量|已售|售出/i.test(line)) || "";
  };
  const usefulCard = (anchor) => {
    let node = anchor;
    for (let i = 0; i < 6 && node && node !== document.body; i += 1) {
      const text = clean(node.innerText);
      if (node.querySelector("img") && (text.includes("$") || text.length > 20)) return node;
      node = node.parentElement;
    }
    return anchor;
  };

  const anchors = Array.from(document.querySelectorAll(
    'a.S-product-card__img-container[href*="-p-"], a.j-expose__product-item-img[href*="-p-"], a[href*="goods_id="]'
  ));
  const byKey = new Map();
  for (const anchor of anchors) {
    const href = absolute(anchor.getAttribute("href"));
    const goodsId = goodsIdFromHref(href);
    if (!href || (!goodsId && !href.includes("shein.com"))) continue;
    const card = usefulCard(anchor);
    const img = anchor.querySelector(
      'img.crop-image-container__img[src*="ltwebstatic"], img.crop-image-container__img[data-src*="ltwebstatic"]'
    );
    const text = clean(card.innerText);
    const prices = priceMatches(text);
    const row = {
      goods_id: goodsId,
      name: titleFrom(anchor, card),
      image: productImage(img),
      sales: salesFromText(text),
      price: prices[0] || "",
      original_price: prices.length > 1 ? prices[prices.length - 1] : "",
      link: href,
    };
    const key = row.goods_id || row.link;
    if (key && (row.name || row.image || row.price)) byKey.set(key, row);
  }
  return {
    url: location.href,
    title: document.title,
    rows: Array.from(byKey.values()),
  };
}
"""


def choose_shein_page(pages: list[Any]) -> Any:
    candidates = [page for page in pages if "shein.com" in page.url and not page.url.startswith("chrome")]
    if not candidates:
        raise RuntimeError("9222 Chrome 里没有打开 shein.com 页面")
    normal = [page for page in candidates if "/risk/action/limit" not in page.url]
    return (normal or candidates)[-1]


def collect_visible_payload(page: Any) -> dict[str, Any]:
    return page.evaluate(EXTRACT_JS)


def merge_payload(base: dict[str, Any], extra: dict[str, Any]) -> dict[str, Any]:
    rows = {str(item.get("goods_id") or item.get("link")): item for item in base["rows"]}
    for item in extra["rows"]:
        key = str(item.get("goods_id") or item.get("link") or "")
        if key:
            rows[key] = item
    base["rows"] = list(rows.values())
    return base


def collect_page(page: Any, do_scroll: bool) -> dict[str, Any]:
    payload = collect_visible_payload(page)
    if not do_scroll:
        return payload
    last_y = -1
    for _ in range(20):
        page.evaluate("window.scrollBy(0, Math.floor(window.innerHeight * 0.85))")
        page.wait_for_timeout(700)
        payload = merge_payload(payload, collect_visible_payload(page))
        current = page.evaluate("Math.round(window.scrollY + window.innerHeight)")
        height = page.evaluate("document.body.scrollHeight")
        if current >= height - 5 or current == last_y:
            break
        last_y = current
    page.evaluate("window.scrollTo(0, 0)")
    return payload


def load_or_create(path: Path) -> Any:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "商品"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {"A": 20, "B": 35, "C": 80, "D": 16, "E": 70, "F": 90, "G": 14, "H": 12, "I": 12, "J": 90}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    return wb


def existing_keys(ws: Any) -> set[str]:
    keys: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        goods_id = str(row[3] or "")
        link = str(row[9] or "")
        if goods_id:
            keys.add(goods_id)
        elif link:
            keys.add(link)
    return keys


def append_rows(output: Path, payload: dict[str, Any]) -> tuple[int, int]:
    wb = load_or_create(output)
    ws = wb["商品"]
    seen = existing_keys(ws)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    added = 0
    for item in payload["rows"]:
        key = str(item.get("goods_id") or item.get("link") or "")
        if not key or key in seen:
            continue
        ws.append(
            [
                now,
                payload["title"],
                payload["url"],
                item.get("goods_id", ""),
                item.get("name", ""),
                item.get("image", ""),
                item.get("sales", ""),
                item.get("price", ""),
                item.get("original_price", ""),
                item.get("link", ""),
            ]
        )
        seen.add(key)
        added += 1
    ws.auto_filter.ref = ws.dimensions
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return added, len(payload["rows"])


def repair_current_page_images(output: Path, payload: dict[str, Any]) -> tuple[int, int]:
    wb = load_workbook(output)
    ws = wb["商品"]
    images_by_id = {
        str(item.get("goods_id") or ""): str(item.get("image") or "")
        for item in payload["rows"]
        if item.get("goods_id") and is_real_product_image(str(item.get("image") or ""))
    }
    updated = 0
    for row in ws.iter_rows(min_row=2):
        goods_id = str(row[3].value or "")
        image = images_by_id.get(goods_id)
        if image and row[5].value != image:
            row[5].value = image
            updated += 1
    wb.save(output)
    return updated, len(images_by_id)


def is_real_product_image(image: str) -> bool:
    return "img.ltwebstatic.com" in image and "bg-grey" not in image


def load_detail_image(context: Any, url: str) -> str:
    for attempt in range(1, 4):
        page = context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_timeout(2500)
            image = str(
                page.evaluate(
                    """
                    () => {
                      const meta = document.querySelector('meta[property="og:image"]');
                      const main = document.querySelector(
                        '.product-intro__main img[data-src], .product-intro__main img[src], img.crop-image-container__img[data-src], img.crop-image-container__img[src]'
                      );
                      return (meta && meta.content) || (main && (main.getAttribute('data-src') || main.getAttribute('src'))) || '';
                    }
                    """
                )
                or ""
            )
            page.close()
            if image.startswith("//"):
                image = "https:" + image
            if is_real_product_image(image):
                return image
        except Exception as exc:
            print(f"  详情页第 {attempt}/3 次打开失败: {exc}")
            try:
                page.close()
            except Exception:
                pass
            time.sleep(3)
    return ""


def load_image_payload(context: Any, url: str) -> dict[str, Any] | None:
    for attempt in range(1, 4):
        page = context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
            page.wait_for_timeout(3000)
            payload = collect_visible_payload(page)
            page.close()
            return payload
        except Exception as exc:
            print(f"  第 {attempt}/3 次打开失败: {exc}")
            try:
                page.close()
            except Exception:
                pass
            time.sleep(3)
    return None


def repair_images(context: Any, output: Path) -> tuple[int, int, int, Path]:
    wb = load_workbook(output)
    ws = wb["商品"]
    page_urls = list(dict.fromkeys(str(row[2] or "") for row in ws.iter_rows(min_row=2, values_only=True) if row[2]))
    images_by_id: dict[str, str] = {}
    failed_pages = 0
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = output.with_name(f"{output.stem}.before_image_repair_{stamp}{output.suffix}")
    shutil.copy2(output, backup)
    for index, page_url in enumerate(page_urls, start=1):
        print(f"修复主图 {index}/{len(page_urls)}: {page_url}")
        payload = load_image_payload(context, page_url)
        if payload is None:
            failed_pages += 1
            print("  跳过该页，继续修复后续分页")
            continue
        for item in payload["rows"]:
            goods_id = str(item.get("goods_id") or "")
            image = str(item.get("image") or "")
            if goods_id and is_real_product_image(image):
                images_by_id[goods_id] = image

    missing = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        goods_id = str(row[3] or "")
        link = str(row[9] or "")
        if goods_id and goods_id not in images_by_id and link:
            missing.append((goods_id, link))
    for index, (goods_id, link) in enumerate(missing, start=1):
        print(f"详情页补主图 {index}/{len(missing)}: goods_id={goods_id}")
        image = load_detail_image(context, link)
        if image:
            images_by_id[goods_id] = image

    updated = 0
    for row in ws.iter_rows(min_row=2):
        goods_id = str(row[3].value or "")
        image = images_by_id.get(goods_id)
        if image and row[5].value != image:
            row[5].value = image
            updated += 1

    wb.save(output)
    return updated, len(images_by_id), failed_pages, backup


def check_cdp(port: int) -> str:
    url = f"http://127.0.0.1:{port}/json/version"
    try:
        with urlopen(url, timeout=3) as response:
            if response.status != 200:
                raise RuntimeError(f"{url} 返回 HTTP {response.status}")
            text = response.read().decode("utf-8", "replace")
            data = json.loads(text)
            ws_url = str(data.get("webSocketDebuggerUrl") or "")
            if not ws_url.startswith("ws://"):
                raise RuntimeError(f"{url} 没有返回可用的 webSocketDebuggerUrl")
            return ws_url
    except URLError as exc:
        raise RuntimeError(
            f"连接不到 Chrome 调试端口 {port}。\n"
            "请先用下面命令启动浏览器：\n"
            r'"C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\Users\ln\AppData\Local\chrome_debug_9222"'
        ) from exc


def connect_browser(playwright: Any, port: int) -> Any:
    ws_url = check_cdp(port)
    last_error: Exception | None = None
    for _ in range(3):
        try:
            return playwright.chromium.connect_over_cdp(ws_url, timeout=15000)
        except Exception as exc:
            last_error = exc
            time.sleep(1)
    raise RuntimeError(
        f"Chrome 调试端口 {port} 可以访问，但 Playwright 连接 CDP 失败。\n"
        "通常是 Chrome 刚启动还没准备好，或旧的调试浏览器卡住了。可以先关闭这个 9222 Chrome 后重新用指定命令启动。\n"
        f"原始错误: {last_error}"
    ) from last_error


def capture_current_page(port: int, output: Path, do_scroll: bool) -> tuple[int, int, str]:
    with sync_playwright() as p:
        browser = connect_browser(p, port)
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = choose_shein_page(context.pages)
        payload = collect_page(page, do_scroll)
    added, found = append_rows(output, payload)
    return added, found, payload["url"]


def main() -> None:
    parser = argparse.ArgumentParser(description="连接 9222 Chrome，抓取当前 SHEIN 页面商品并追加到 Excel。")
    parser.add_argument("--port", type=int, default=9222)
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--no-scroll", action="store_true", help="不自动滚动当前页，只抓当前已渲染内容")
    parser.add_argument("--repair-images", action="store_true", help="重访 Excel 记录的分页，按 goods_id 修复主图列")
    parser.add_argument("--repair-current-page-images", action="store_true", help="只读取当前已打开页面，修复该页商品主图")
    args = parser.parse_args()

    if args.repair_images:
        raise RuntimeError("自动跳页修复会触发 SHEIN 风控，已禁用。请手动翻页后使用 --repair-current-page-images。")

    if args.repair_current_page_images:
        output = Path(args.output)
        if not output.exists():
            raise RuntimeError(f"待修复 Excel 不存在: {output}")
        with sync_playwright() as p:
            browser = connect_browser(p, args.port)
            context = browser.contexts[0] if browser.contexts else browser.new_context()
            page = choose_shein_page(context.pages)
            payload = collect_visible_payload(page)
        updated, found = repair_current_page_images(output, payload)
        print(f"当前页: {payload['url']}")
        print(f"当前页识别真实商品主图: {found}; 已更新: {updated}")
        print(f"修复后 Excel: {output}")
        return

    added, found, url = capture_current_page(args.port, Path(args.output), not args.no_scroll)
    print(f"当前页: {url}")
    print(f"本页识别商品: {found}; 新增写入: {added}")
    if found == 0:
        print("提示: 当前页没有识别到商品卡。请确认页面商品列表已正常显示，不是空列表/风控页/只显示店铺推荐卡。")
    print(f"Excel: {Path(args.output)}")


if __name__ == "__main__":
    try:
        main()
    except RuntimeError as exc:
        print(f"错误: {exc}")
        raise SystemExit(1) from None
